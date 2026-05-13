# backend/app/agents/run_batch_test.py
# Exécute le pipeline sur toutes les questions de questions.json
# et exporte les résultats dans un fichier Excel.

import json
import time
from datetime import datetime
from pathlib import Path

import pandas as pd

from app.agents.orchestrator import build_initial_state, run_pipeline

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

QUESTIONS_FILE = Path(__file__).parent / "questions.json"
OUTPUT_DIR = Path(__file__).parent
USER_ROLE = "ADMIN"
USER_ID = "batch-test"
USER_SHOP_ID = None  # None pour ADMIN

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_response(formatted_response: dict | None) -> str:
    """Extrait la réponse finale lisible depuis formatted_response."""
    if not formatted_response:
        return ""
    ftype = formatted_response.get("type", "text")
    if ftype == "text":
        return formatted_response.get("content", "")
    if ftype == "table":
        cols = [c["label"] for c in formatted_response.get("columns", [])]
        rows = formatted_response.get("data", [])
        lines = [" | ".join(cols)]
        lines.append("-" * len(lines[0]))
        for row in rows[:20]:  # max 20 lignes dans la cellule
            lines.append(" | ".join(str(row.get(c["field"], "")) for c in formatted_response["columns"]))
        if len(rows) > 20:
            lines.append(f"... ({len(rows)} lignes au total)")
        return "\n".join(lines)
    if ftype == "chart":
        x = formatted_response.get("xAxis", {}).get("label", "")
        y = formatted_response.get("yAxis", {}).get("label", "")
        chart_type = formatted_response.get("chartType", "bar")
        data = formatted_response.get("data", [])
        lines = [f"[Graphique {chart_type} — X: {x} | Y: {y}]"]
        x_field = formatted_response["xAxis"]["field"]
        y_field = formatted_response["yAxis"]["field"]
        for row in data[:20]:
            lines.append(f"  {row.get(x_field)} → {row.get(y_field)}")
        return "\n".join(lines)
    return str(formatted_response)


# ---------------------------------------------------------------------------
# Main batch run
# ---------------------------------------------------------------------------

def run_batch():
    questions = json.loads(QUESTIONS_FILE.read_text(encoding="utf-8"))
    total = len(questions)
    print(f"\n{'='*65}")
    print(f"  BATCH TEST — {total} questions | rôle: {USER_ROLE}")
    print(f"{'='*65}\n")

    records = []

    for i, question in enumerate(questions, 1):
        print(f"[{i:02d}/{total}] {question[:70]}")
        t0 = time.perf_counter()

        try:
            state = build_initial_state(
                question=question,
                user_id=USER_ID,
                user_role=USER_ROLE,
                user_shop_id=USER_SHOP_ID,
            )
            result = run_pipeline(state)
            elapsed = round(time.perf_counter() - t0, 2)

            intention    = result.get("intention") or ""
            rewritten    = result.get("rewritten_query") or ""
            sql          = result.get("sql_query") or ""
            exec_error   = result.get("execution_error") or ""
            pipeline_err = result.get("error") or ""
            fr           = result.get("formatted_response")
            final_resp   = _extract_response(fr)
            fmt_type     = (fr or {}).get("type", "")

            status = "✅" if not exec_error and not pipeline_err else "❌"
            print(f"       {status} intention={intention} | format={fmt_type} | {elapsed}s")

        except Exception as e:
            elapsed = round(time.perf_counter() - t0, 2)
            intention = rewritten = sql = final_resp = fmt_type = ""
            exec_error = pipeline_err = str(e)
            print(f"       ❌ EXCEPTION: {e}")

        records.append({
            "N°":                i,
            "Question":          question,
            "Query réécrite":    rewritten,
            "Intention":         intention,
            "SQL généré":        sql,
            "Erreur exécution":  exec_error,
            "Erreur pipeline":   pipeline_err,
            "Type format":       fmt_type,
            "Réponse finale":    final_resp,
            "Durée (s)":         elapsed,
        })

    # ---------------------------------------------------------------------------
    # Export Excel
    # ---------------------------------------------------------------------------

    df = pd.DataFrame(records)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"batch_results_{timestamp}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Résultats")

        ws = writer.sheets["Résultats"]

        # Largeurs de colonnes
        col_widths = {
            "A": 5,   # N°
            "B": 55,  # Question
            "C": 55,  # Query réécrite
            "D": 12,  # Intention
            "E": 70,  # SQL généré
            "F": 40,  # Erreur exécution
            "G": 40,  # Erreur pipeline
            "H": 12,  # Type format
            "I": 80,  # Réponse finale
            "J": 10,  # Durée
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Wrap text pour SQL et réponse finale
        from openpyxl.styles import Alignment, PatternFill, Font
        wrap = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

        # Header style
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Colorier les lignes en erreur en rouge clair
        error_fill = PatternFill("solid", fgColor="FFD7D7")
        ok_fill    = PatternFill("solid", fgColor="E8F5E9")
        for row in ws.iter_rows(min_row=2):
            has_error = any(row[5].value or row[6].value, )
            fill = error_fill if has_error else ok_fill
            for cell in row:
                if cell.fill.fgColor.rgb == "00000000":  # pas déjà colorié
                    cell.fill = fill

        ws.freeze_panes = "A2"

    success = sum(1 for r in records if not r["Erreur exécution"] and not r["Erreur pipeline"])
    print(f"\n{'='*65}")
    print(f"  ✅ {success}/{total} succès")
    print(f"  📁 Résultats exportés : {output_path}")
    print(f"{'='*65}\n")
    return str(output_path)


if __name__ == "__main__":
    run_batch()
    # Flush Opik traces so all spans are uploaded before the script exits
    try:
        from opik import get_global_client
        get_global_client().flush()
    except Exception:
        pass
